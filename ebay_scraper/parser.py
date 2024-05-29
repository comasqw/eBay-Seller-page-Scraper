import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
import json
import openpyxl
import os
from datetime import datetime

ua = UserAgent()

headers = {
    "User-Agent": ua.random,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
}

class Parser:
    def __init__(self, url):
        """
        Initializes an instance of the Parser class.

        Arguments:
        url (str) - Initially, this is the URL of the seller's page, but later it will become the URL of all the seller's products.

        Attributes:
        self.url (str) - The URL used for sending requests. Initially, it's the seller's page URL, but it will later change to the URL of the seller's products page.
        self.headers (dict) - Headers to send with the GET request.
        self.params (dict) - Parameters to send with the GET request, such as filters.
        self.information_dct (dict or None) - Initially None, but later it will become a dictionary containing information about the seller, including their name, feedback, items sold, categories, etc.
        self.file_format (str or None) - Initially None, but later the user will choose the desired file format - "Json" or "Excel".
        """
        self.url = url
        self.headers = headers
        self.params = {}
        self.information_dct = None
        self.file_format = None
    def get_request(self):
        """
        Sends a GET request to the specified URL with the given headers and parameters.

        :return: Returns the response object obtained from the GET request.
        """
        return requests.get(self.url, headers=headers, params=self.params)

    def get_html(self):
        """
        Sends a GET request to the specified URL and returns the HTML content of the page.

        :return: Returns the HTML content of the page obtained from the GET request.
        """
        html = self.get_request().text
        return html


    def make_soup_object(self):
        """
        Sends a GET request to the specified URL and creates a BeautifulSoup object from the HTML content.

        :return: Creates and returns a BeautifulSoup object.
        """
        html = self.get_html()
        return BeautifulSoup(html, "lxml")

    def change_url_to_seller_all_products(self):
        soup = self.make_soup_object()
        link = soup.find("a", class_="str-marginals__footer--button")
        self.url = link.get("href").replace("&amp;", "&")
        self.params["_ipg"] = 240

    def get_information(self):

        soup = self.make_soup_object()

        categories_lst = []
        for category in soup.find("ul", class_="srp-refine__category__list").find_all("li"):
            link = category.find("a").get("href")
            categories_lst.append([category.text, link.split("&")])

        categories_dct = {"All": "0"}

        for category in categories_lst[1:]:
            categories_dct[category[0]] = category[-1][-1].split("=")[-1]

        information_lst = [i.text for i in soup.find(class_="str-seller-card__stats-content").find_all("div")]
        information_dct = {}
        information_dct["Seller"] = soup.find(class_="str-seller-card__store-name").text
        information_dct["Feedback"] = information_lst[0]
        information_dct["Items Sold"] = information_lst[1]
        information_dct["Followers"] = information_lst[2]
        information_dct["Items Count"] = soup.find(class_="srp-controls__count-heading").text
        information_dct["Categories"] = categories_dct

        self.information_dct = information_dct

    def make_categories_lst(self):
        categories_lst = [category for category in self.information_dct["Categories"]]
        return categories_lst

    def get_items_in_category(self):
        soup = self.make_soup_object()
        self.information_dct["Items Count"] = soup.find(class_="srp-controls__count-heading").text

    def choose_category(self, category):
        self.url = self.url.replace("&store_cat=0", "")
        categories_dct = self.information_dct["Categories"]
        self.params["store_cat"] = categories_dct[category]

    def set_filters(self, filter, value):
        self.params[filter] = value

    def set_file_format(self, file_format):
        self.file_format = file_format

    def get_pages_to_scrape(self):
        items_count = []
        for i in self.information_dct["Items Count"]:
            if i.isnumeric():
                items_count.append(i)
        return int("".join(items_count)) // 240 + 1

    def get_cards(self, pages_to_scrape):

        cards_lst = []
        for page in range(1, pages_to_scrape + 1):
            self.params["_pgn"] = page
            soup = self.make_soup_object()
            cards = soup.find(class_="srp-results srp-list clearfix").find_all(class_="s-item s-item__pl-on-bottom")
            for card in cards:
                card_dct = {}
                title = card.find(class_="s-item__link").text
                link = card.find(class_="s-item__link").get("href")
                price = card.find(class_="s-item__price").text
                logistic_cost = card.find(class_="s-item__shipping s-item__logisticsCost").text
                location = card.find(class_="s-item__location s-item__itemLocation").text
                img_url = card.find("div", class_="s-item__image-wrapper image-treatment").find("img").get("src")

                card_dct["Title"] = title
                card_dct["Price"] = price
                card_dct["Logistic Cost"] = logistic_cost
                card_dct["Location"] = location
                card_dct["Link"] = link
                card_dct["Image"] = img_url

                cards_lst.append(card_dct)
            print(f"Page {page} scraped!")
        return cards_lst

    def save_cards(self):
        os.chdir("results")
        cards_lst = self.get_cards(self.get_pages_to_scrape())
        if self.file_format == "json":
            with open(f"{self.information_dct['Seller']}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.json", "w") as file:
                json.dump(cards_lst, file, indent=4)
        elif self.file_format == "excel":
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1, column=1, value="Title")
            sheet.cell(row=1, column=2, value="Price")
            sheet.cell(row=1, column=3, value="Logistic Cost")
            sheet.cell(row=1, column=4, value="Location")
            sheet.cell(row=1, column=5, value="Link")
            sheet.cell(row=1, column=6, value="Image")

            row = 2
            for card in cards_lst:
                sheet.cell(row=row, column=1, value=card["Title"])
                sheet.cell(row=row, column=2, value=card["Price"])
                sheet.cell(row=row, column=3, value=card["Logistic Cost"])
                sheet.cell(row=row, column=4, value=card["Location"])
                sheet.cell(row=row, column=5, value=card["Link"])
                sheet.cell(row=row, column=6, value=card["Image"])

                row += 1

            wb.save(f"{self.information_dct['Seller']}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
            wb.close()
        os.chdir("..")